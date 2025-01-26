from ui import Win
from tkinter import filedialog
from openpyxl import Workbook
from tkinter import messagebox

class Controller:
	# 导入UI类后，替换以下的 object 类型，将获得 IDE 属性提示功能
	ui: Win
	tableData: []
	fields = ('商品名称', '单位', '单价（元）', '数量', '金额（元）')

	def __init__(self):
		pass
	def init(self, ui):
		"""
		得到UI实例，对组件进行初始化配置
		"""
		self.ui = ui
		# TODO 组件初始化 赋值操作
		self.tableData = []

	class MyErr(Exception):
		pass

	# 定义异常处理装饰器
	def handle_exceptions(func):
		def wrapper(self, *args, **kwargs):
			try:
				return func(self, *args, **kwargs)
			except Controller.MyErr as e:
				messagebox.showwarning('警告', e)
			except Exception as e:
				messagebox.showerror('未知错误', e)
		return wrapper


	def get_data(self):
		product = self.ui.tk_input_product.get()
		unit = self.ui.tk_input_unit.get()
		price = self.ui.tk_input_price.get()
		num = self.ui.tk_input_num.get()
		sum = self.ui.tk_input_sum.get()
		return product, unit, price, num, sum
	def get_input(self):
		item = self.get_data()
		if '' in item:
			idx = item.index('')
			arr = self.fields
			raise self.MyErr(f'{arr[idx]}不能为空')
		return item

	def get_selection(self):
		table = self.ui.tk_table_
		selection = table.selection()
		if not selection:
			raise self.MyErr('没有选中项')
		return table.item(selection[0], 'values')

	def clear_data(self):
		self.ui.tk_input_product.delete(0, 'end')
		self.ui.tk_input_unit.delete(0, 'end')
		self.ui.tk_input_price.delete(0, 'end')
		self.ui.tk_input_num.delete(0, 'end')
		self.ui.tk_input_sum.delete(0, 'end')
	def sync(self):
		table = self.ui.tk_table_
		tableData = self.tableData
		for item in table.get_children():
			table.delete(item)
		for item in tableData:
			table.insert('', 'end', values=item)

	def get_names(self):
		text = self.ui.tk_text_names
		names = (text.get("0.0", "end").replace(" ", "")).split("\n")
		names = [x for x in names if x != '' and x is not None]
		if len(names) < 1:
			raise self.MyErr('请输入人名')
		return names
	def get_output_path(self):
		output_path = filedialog.asksaveasfilename(
			title="保存为",
			defaultextension=".xlsx",  # 默认文件扩展名
			filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]  # 文件类型过滤器
		)
		if not output_path:
			raise self.MyErr('请选择文件存储路径')
		return output_path





	@handle_exceptions
	def add(self,evt):
		item = self.get_input()
		self.tableData.append(item)
		self.sync()
		self.clear_data()

	@handle_exceptions
	def delete(self,evt):
		item = self.get_selection()
		self.tableData.remove(item)
		self.sync()

	@handle_exceptions
	def modify(self,evt):
		old = self.get_selection()
		new_item = self.get_input()
		self.tableData[self.tableData.index(old)] = new_item
		self.sync()
		self.clear_data()

	@handle_exceptions
	def up(self,evt):
		item = self.get_selection()
		x = self.tableData.index(item)
		lst = self.tableData
		if x > 0:
			lst[x - 1], lst[x] = lst[x], lst[x - 1]
			self.tableData = lst
			self.sync()
	@handle_exceptions
	def down(self,evt):
		item = self.get_selection()
		x = self.tableData.index(item)
		lst = self.tableData
		if x < len(lst):
			lst[x], lst[x+1] = lst[x+1], lst[x]
			self.tableData = lst
			self.sync()

	@handle_exceptions
	def generate(self,evt):
		tableData = self.tableData
		if (len(tableData) < 1):
			raise self.MyErr('表格不能为空')

		names = self.get_names()
		output_path = self.get_output_path()

		self.generateExcel(tableData, names, output_path)

	def generateExcel(self, products, names, filepath):

		result = []
		for idx, name in enumerate(names):
			for product in products:
				result.append([idx + 1, *product, name])

		# 创建一个新的工作簿和工作表
		wb = Workbook()
		ws = wb.active

		# 定义表头
		headers = ['序号', *self.fields, '人名']
		# 写入表头
		ws.append(headers)

		# 开始行索引，从第 2 行开始写入数据
		start_row = 2

		# 处理合并单元格
		for col in ['序号', '人名']:
			col_index = headers.index(col) + 1
			i = 0
			while i < len(result):
				j = i + 1
				# 找到连续相同值的范围
				while j < len(result) and result[j][col_index - 1] == result[i][col_index - 1]:
					j += 1
				if j > i + 1:
					# 合并单元格
					ws.merge_cells(start_row=start_row + i, start_column=col_index,
								   end_row=start_row + j - 1, end_column=col_index)
				# 将值写入合并区域的起始单元格
				ws.cell(row=start_row + i, column=col_index, value=result[i][col_index - 1])
				i = j

		# 写入其他列的数据
		for row_num, row_data in enumerate(result, start=start_row):
			for col_index, value in enumerate(row_data, start=1):
				if headers[col_index - 1] not in ['序号', '人名']:
					ws.cell(row=row_num, column=col_index, value=value)

		# 保存文件
		wb.save(filepath)
		messagebox.showinfo('成功', "导出完成")
